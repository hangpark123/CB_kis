import openpyxl
import os

def find_api_in_excel(filename, keywords):
    print(f"Loading {filename}...")
    try:
        wb = openpyxl.load_workbook(filename, read_only=True, data_only=True)
        print(f"Sheets: {wb.sheetnames}")
        
        found_count = 0
        
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            print(f"\n--- Scanning Sheet: {sheet_name} ---")
            
            # 헤더 읽기 (첫 5줄만)
            for i, row in enumerate(ws.iter_rows(max_row=500, values_only=True)):
                row_str = " ".join([str(c) for c in row if c])
                
                # 키워드 검색
                if any(k in row_str for k in keywords):
                    print(f"Found in row {i+1}: {row_str[:200]}...")
                    found_count += 1
                    if found_count > 10:  # 너무 많이 나오면 중단
                        print("... too many results in this sheet")
                        break
                        
    except Exception as e:
        print(f"Error: {e}")

if __name__ == "__main__":
    import sys
    sys.stdout.reconfigure(encoding='utf-8')
    
    filename = "한국투자증권_오픈API_전체문서_20260108_030000.xlsx"
    if not os.path.exists(filename):
        print("File not found.")
    else:
        find_api_in_excel(filename, ["종목", "검색", "Search", "Master", "CTPF", "JTTT"])
